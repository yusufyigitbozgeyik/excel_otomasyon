from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

STOK_HEADERS = [
    "Stok Kodu", "Ürün Adı", "Kategori", "Miktar", "Birim", "Kritik Seviye", "Açıklama", "Son Güncelleme"
]

# --- Temel Fonksiyonlar ---
def create_stock_file(filename):
    """Yeni bir stok Excel dosyası oluşturur."""
    wb = Workbook()
    ws = wb.active
    ws.append(STOK_HEADERS)
    wb.save(filename)


def file_exists(filename):
    return os.path.exists(filename)


def read_all_products(filename):
    """Tüm ürünleri (başlık hariç) listeler."""
    wb = load_workbook(filename)
    ws = wb.active
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # başlık
        data.append(list(row))
    return data


def add_product(filename, product):
    """Yeni ürün ekler. product: [Stok Kodu, Ürün Adı, Kategori, Miktar, Birim, Kritik Seviye, Açıklama]"""
    wb = load_workbook(filename)
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append(product + [now])
    wb.save(filename)


def update_product(filename, row_index, new_product):
    """Belirtilen satırı yeni ürün verileriyle günceller. (row_index: 0 tabanlı, başlık hariç)"""
    wb = load_workbook(filename)
    ws = wb.active
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for col, value in enumerate(new_product + [now], start=1):
        ws.cell(row=row_index+2, column=col, value=value)
    wb.save(filename)


def delete_product(filename, row_index):
    """Belirtilen satırı siler. (row_index: 0 tabanlı, başlık hariç)"""
    wb = load_workbook(filename)
    ws = wb.active
    ws.delete_rows(row_index+2)
    wb.save(filename)


def stock_in_out(filename, row_index, amount):
    """Stok giriş/çıkış işlemi. amount pozitifse giriş, negatifse çıkış."""
    wb = load_workbook(filename)
    ws = wb.active
    miktar = ws.cell(row=row_index+2, column=4).value or 0
    yeni_miktar = float(miktar) + float(amount)
    ws.cell(row=row_index+2, column=4, value=yeni_miktar)
    ws.cell(row=row_index+2, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    wb.save(filename)


def get_critical_products(filename):
    """Kritik seviyedeki ürünleri döndürür."""
    products = read_all_products(filename)
    critical = []
    for p in products:
        try:
            miktar = float(p[3])
            kritik = float(p[5])
            if miktar <= kritik:
                critical.append(p)
        except:
            continue
    return critical


def filter_products(filename, keyword=None, category=None):
    """Ürünleri arama ve kategoriye göre filtreler."""
    products = read_all_products(filename)
    result = []
    for p in products:
        if keyword and keyword.lower() not in str(p[1]).lower():
            continue
        if category and category.lower() != str(p[2]).lower():
            continue
        result.append(p)
    return result


def get_categories(filename):
    """Tüm kategorileri benzersiz olarak döndürür."""
    products = read_all_products(filename)
    return sorted(set([p[2] for p in products if p[2]]))


def backup_file(filename):
    """Excel dosyasının yedeğini oluşturur."""
    import shutil
    backup_name = filename.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    shutil.copy(filename, backup_name)
    return backup_name


def get_summary(filename):
    """Toplam ürün, toplam stok, kritik ürün sayısı gibi özet rapor döndürür."""
    products = read_all_products(filename)
    toplam_urun = len(products)
    toplam_stok = sum([float(p[3]) for p in products if p[3]])
    kritik_urun = len(get_critical_products(filename))
    return {
        "toplam_urun": toplam_urun,
        "toplam_stok": toplam_stok,
        "kritik_urun": kritik_urun
    } 